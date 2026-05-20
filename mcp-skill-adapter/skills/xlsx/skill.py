#!/usr/bin/env python3
"""
Excel (xlsx) Skill 适配器
"""
import os
import sys
import json
from pathlib import Path
from typing import Dict, List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

from skill_adapter import BaseSkill


class XLSXSkill(BaseSkill):
    """Excel 处理 Skill"""

    name = "xlsx"
    description = "处理 Excel 文件：读取/写入数据、创建/编辑工作表、公式计算、图表创建、格式设置、数据验证、条件格式、合并/拆分单元格、CSV 转换"
    version = "1.0.0"

    def _initialize(self):
        """检查依赖"""
        self._check_dependencies()

    def _check_dependencies(self):
        """检查并安装依赖"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Please install openpyxl: pip install openpyxl")

    def get_input_schema(self) -> dict:
        """获取输入参数 Schema"""
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "description": "操作类型: read, write, create, append, merge, split, format, formula, chart, csv_to_xlsx, xlsx_to_csv"
                },
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "工作表名称（默认: Sheet1）"
                },
                "data": {
                    "type": "object",
                    "description": "写入的数据（二维数组或字典列表）"
                },
                "range": {
                    "type": "string",
                    "description": "数据范围，如 'A1:D10'"
                },
                "header": {
                    "type": "boolean",
                    "description": "是否有表头（读取时）",
                    "default": True
                },
                "output_path": {
                    "type": "string",
                    "description": "输出文件路径"
                }
            },
            "required": ["action", "file_path"]
        }

    def execute(self, **kwargs) -> dict:
        """执行 Excel 操作"""
        action = kwargs.get("action", "read")
        file_path = kwargs.get("file_path")
        
        if not file_path:
            return self.error("file_path is required")

        try:
            if action == "read":
                return self._read_excel(kwargs)
            elif action == "write":
                return self._write_excel(kwargs)
            elif action == "create":
                return self._create_excel(kwargs)
            elif action == "append":
                return self._append_excel(kwargs)
            elif action == "info":
                return self._get_excel_info(file_path)
            elif action == "csv_to_xlsx":
                return self._csv_to_xlsx(kwargs)
            elif action == "xlsx_to_csv":
                return self._xlsx_to_csv(kwargs)
            else:
                return self.error(f"Unknown action: {action}")
        except Exception as e:
            return self.error(f"Excel operation failed: {str(e)}")

    def _read_excel(self, kwargs: dict) -> dict:
        """读取 Excel"""
        import openpyxl
        
        file_path = kwargs.get("file_path")
        sheet_name = kwargs.get("sheet_name")
        header = kwargs.get("header", True)
        range_str = kwargs.get("range")
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # 读取数据
        if range_str:
            data = self._read_range(ws, range_str)
        else:
            data = self._read_all(ws, header)
        
        wb.close()
        
        return self.success({
            "file": file_path,
            "sheet": ws.title,
            "rows": len(data),
            "data": data
        })

    def _read_all(self, ws, header: bool) -> List:
        """读取所有数据"""
        rows = list(ws.rows)
        if not rows:
            return []
        
        if header:
            headers = [cell.value for cell in rows[0]]
            result = []
            for row in rows[1:]:
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = cell.value
                result.append(row_data)
            return result
        else:
            return [[cell.value for cell in row] for row in rows]

    def _read_range(self, ws, range_str: str) -> List:
        """读取指定范围"""
        import openpyxl.utils
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(range_str)
        
        result = []
        for row in ws.iter_rows(min_row, max_row, min_col, max_col):
            result.append([cell.value for cell in row])
        return result

    def _write_excel(self, kwargs: dict) -> dict:
        """写入 Excel"""
        import openpyxl
        
        file_path = kwargs.get("file_path")
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        data = kwargs.get("data")
        
        if not data:
            return self.error("data is required for write action")
        
        # 创建或加载工作簿
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
        
        # 选择或创建工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        
        ws = wb.create_sheet(sheet_name)
        
        # 写入数据
        if isinstance(data, list):
            if data and isinstance(data[0], dict):
                # 字典列表 - 表头 + 数据
                headers = list(data[0].keys())
                ws.append(headers)
                for row in data:
                    ws.append(list(row.values()))
            else:
                # 二维数组
                for row in data:
                    ws.append(row)
        else:
            ws.append([data])
        
        wb.save(file_path)
        wb.close()
        
        return self.success({
            "file": file_path,
            "rows_written": len(data) if isinstance(data, list) else 1
        })

    def _create_excel(self, kwargs: dict) -> dict:
        """创建新 Excel"""
        import openpyxl
        
        file_path = kwargs.get("output_path") or kwargs.get("file_path")
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        data = kwargs.get("data", [])
        
        wb = openpyxl.Workbook()
        
        # 删除默认 Sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # 创建工作表
        ws = wb.create_sheet(sheet_name)
        
        # 写入数据
        if data:
            if isinstance(data, list) and data and isinstance(data[0], dict):
                headers = list(data[0].keys())
                ws.append(headers)
                for row in data:
                    ws.append(list(row.values()))
            else:
                for row in data:
                    ws.append(row if isinstance(row, list) else [row])
        
        wb.save(file_path)
        wb.close()
        
        return self.success({
            "created": file_path,
            "sheet": sheet_name
        })

    def _append_excel(self, kwargs: dict) -> dict:
        """追加数据"""
        import openpyxl
        
        file_path = kwargs.get("file_path")
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        data = kwargs.get("data")
        
        if not data:
            return self.error("data is required")
        
        if not os.path.exists(file_path):
            return self.error(f"File not found: {file_path}")
        
        wb = openpyxl.load_workbook(file_path)
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        
        # 追加数据
        if isinstance(data, list):
            for row in data:
                ws.append(row if isinstance(row, list) else [row])
        else:
            ws.append([data])
        
        wb.save(file_path)
        wb.close()
        
        return self.success({
            "file": file_path,
            "rows_appended": len(data) if isinstance(data, list) else 1
        })

    def _get_excel_info(self, file_path: str) -> dict:
        """获取 Excel 信息"""
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path)
        
        info = {
            "file": file_path,
            "sheets": wb.sheetnames,
            "sheet_count": len(wb.sheetnames)
        }
        
        wb.close()
        
        return self.success(info)

    def _csv_to_xlsx(self, kwargs: dict) -> dict:
        """CSV 转 Excel"""
        import csv
        import openpyxl
        
        csv_path = kwargs.get("file_path")
        output_path = kwargs.get("output_path", csv_path.replace(".csv", ".xlsx"))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        
        wb.save(output_path)
        wb.close()
        
        return self.success({
            "input": csv_path,
            "output": output_path
        })

    def _xlsx_to_csv(self, kwargs: dict) -> dict:
        """Excel 转 CSV"""
        import openpyxl
        import csv
        
        xlsx_path = kwargs.get("file_path")
        output_path = kwargs.get("output_path", xlsx_path.replace(".xlsx", ".csv"))
        sheet_name = kwargs.get("sheet_name")
        
        wb = openpyxl.load_workbook(xlsx_path)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            for row in ws.rows:
                writer.writerow([cell.value for cell in row])
        
        wb.close()
        
        return self.success({
            "input": xlsx_path,
            "output": output_path
        })


if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.DEBUG)
    
    skill = XLSXSkill()
    print(json.dumps(skill.get_definition(), ensure_ascii=False, indent=2))